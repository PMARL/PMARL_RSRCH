#!/usr/bin/env python3
import argparse
import csv
import json
import math
import statistics
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


FIELDNAMES = [
    "dataset",
    "instance_index",
    "depot",
    "budget_miles",
    "algorithm",
    "collected_prize_excluding_depot",
    "java_raw_prize_including_depot",
    "route_distance_miles",
    "runtime_ms",
    "route",
    "raw_attention_route_distance_miles",
    "repair_applied",
    "feasible_under_java_miles",
]

SUMMARY_FIELDS = [
    "budget_miles",
    "algorithm",
    "n",
    "mean_prize",
    "std_prize",
    "mean_distance_miles",
    "std_distance_miles",
    "mean_runtime_ms",
    "std_runtime_ms",
]

ALGORITHM_ORDER = ["Attention", "P-MARL", "Greedy 2", "Greedy 1"]
COLORS = {
    "Attention": "#2563eb",
    "P-MARL": "#dc2626",
    "Greedy 2": "#16a34a",
    "Greedy 1": "#6b7280",
}


def build_parser():
    root = Path(__file__).resolve().parents[1]
    parser = argparse.ArgumentParser(description="Generate comparison deliverables from raw experiment CSV files.")
    parser.add_argument("--root", default=str(root))
    parser.add_argument("--java-csv", default=str(root / "raw" / "java_baselines.csv"))
    parser.add_argument("--attention-csv", default=str(root / "raw" / "attention_results.csv"))
    return parser


def read_rows(path):
    rows = []
    with open(path, newline="", encoding="utf-8") as handle:
        for row in csv.DictReader(handle):
            normalized = {field: row.get(field, "") for field in FIELDNAMES}
            normalized["instance_index"] = int(normalized["instance_index"])
            normalized["budget_miles"] = int(float(normalized["budget_miles"]))
            normalized["collected_prize_excluding_depot"] = float(normalized["collected_prize_excluding_depot"])
            normalized["route_distance_miles"] = float(normalized["route_distance_miles"])
            normalized["runtime_ms"] = float(normalized["runtime_ms"])
            rows.append(normalized)
    return rows


def write_csv(path, rows, fieldnames):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})


def mean(values):
    return statistics.mean(values) if values else 0.0


def stdev(values):
    return statistics.stdev(values) if len(values) > 1 else 0.0


def summarize(rows):
    grouped = defaultdict(list)
    for row in rows:
        grouped[(row["budget_miles"], row["algorithm"])].append(row)

    summary = []
    for (budget, algorithm), group in sorted(grouped.items(), key=lambda item: (item[0][0], algorithm_rank(item[0][1]))):
        prizes = [row["collected_prize_excluding_depot"] for row in group]
        distances = [row["route_distance_miles"] for row in group]
        runtimes = [row["runtime_ms"] for row in group]
        summary.append({
            "budget_miles": budget,
            "algorithm": algorithm,
            "n": len(group),
            "mean_prize": round(mean(prizes), 3),
            "std_prize": round(stdev(prizes), 3),
            "mean_distance_miles": round(mean(distances), 3),
            "std_distance_miles": round(stdev(distances), 3),
            "mean_runtime_ms": round(mean(runtimes), 3),
            "std_runtime_ms": round(stdev(runtimes), 3),
        })
    return summary


def algorithm_rank(name):
    return ALGORITHM_ORDER.index(name) if name in ALGORITHM_ORDER else len(ALGORITHM_ORDER)


def autosize(ws):
    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 80))
        ws.column_dimensions[letter].width = max(10, min(max_len + 2, 60))


def add_table_sheet(wb, title, rows, fieldnames):
    ws = wb.create_sheet(title)
    ws.append(fieldnames)
    for row in rows:
        ws.append([row.get(field, "") for field in fieldnames])
    header_fill = PatternFill("solid", fgColor="1F2937")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    autosize(ws)
    return ws


def write_workbook(path, raw_rows, summary_rows, manifest):
    wb = Workbook()
    wb.remove(wb.active)
    add_table_sheet(wb, "Summary by Budget", summary_rows, SUMMARY_FIELDS)
    add_table_sheet(wb, "Raw Results", raw_rows, FIELDNAMES)
    ws = wb.create_sheet("Manifest")
    ws.append(["Field", "Value"])
    for key, value in manifest.items():
        ws.append([key, json.dumps(value) if isinstance(value, (list, dict)) else value])
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="1F2937")
    autosize(ws)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def svg_escape(value):
    return str(value).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def plot_svg(path, summary_rows, metric, title, y_label):
    budgets = sorted({int(row["budget_miles"]) for row in summary_rows})
    algorithms = [alg for alg in ALGORITHM_ORDER if any(row["algorithm"] == alg for row in summary_rows)]
    lookup = {(row["budget_miles"], row["algorithm"]): row for row in summary_rows}
    width, height = 920, 560
    left, right, top, bottom = 88, 34, 56, 72
    plot_w = width - left - right
    plot_h = height - top - bottom

    values = []
    for row in summary_rows:
        values.append(float(row[metric]))
        std_key = metric.replace("mean_", "std_")
        if std_key in row:
            values.append(float(row[metric]) + float(row[std_key]))
    y_min = 0.0
    y_max = max(values) if values else 1.0
    if y_max <= 0:
        y_max = 1.0
    y_max *= 1.08

    def x_for(index):
        if len(budgets) == 1:
            return left + plot_w / 2.0
        return left + index * plot_w / (len(budgets) - 1)

    def y_for(value):
        return top + plot_h - (float(value) - y_min) / (y_max - y_min) * plot_h

    lines = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
        '<rect width="100%" height="100%" fill="#ffffff"/>',
        f'<text x="{left}" y="32" font-family="Arial" font-size="22" font-weight="700" fill="#111827">{svg_escape(title)}</text>',
        f'<text x="{left}" y="{height - 18}" font-family="Arial" font-size="13" fill="#374151">Budget (miles)</text>',
        f'<text x="18" y="{top + plot_h / 2}" transform="rotate(-90 18 {top + plot_h / 2})" font-family="Arial" font-size="13" fill="#374151">{svg_escape(y_label)}</text>',
        f'<line x1="{left}" y1="{top}" x2="{left}" y2="{top + plot_h}" stroke="#111827" stroke-width="1.2"/>',
        f'<line x1="{left}" y1="{top + plot_h}" x2="{left + plot_w}" y2="{top + plot_h}" stroke="#111827" stroke-width="1.2"/>',
    ]

    for tick in range(6):
        value = y_min + (y_max - y_min) * tick / 5.0
        y = y_for(value)
        lines.append(f'<line x1="{left}" y1="{y:.2f}" x2="{left + plot_w}" y2="{y:.2f}" stroke="#e5e7eb" stroke-width="1"/>')
        lines.append(f'<text x="{left - 10}" y="{y + 4:.2f}" text-anchor="end" font-family="Arial" font-size="12" fill="#4b5563">{value:.0f}</text>')

    for i, budget in enumerate(budgets):
        x = x_for(i)
        lines.append(f'<line x1="{x:.2f}" y1="{top + plot_h}" x2="{x:.2f}" y2="{top + plot_h + 5}" stroke="#111827"/>')
        lines.append(f'<text x="{x:.2f}" y="{top + plot_h + 24}" text-anchor="middle" font-family="Arial" font-size="12" fill="#111827">{budget}</text>')

    legend_x = left + plot_w - 360
    for i, algorithm in enumerate(algorithms):
        x = legend_x + i * 92
        color = COLORS.get(algorithm, "#111827")
        lines.append(f'<line x1="{x}" y1="30" x2="{x + 18}" y2="30" stroke="{color}" stroke-width="3"/>')
        lines.append(f'<circle cx="{x + 9}" cy="30" r="4" fill="{color}"/>')
        lines.append(f'<text x="{x + 24}" y="34" font-family="Arial" font-size="12" fill="#111827">{svg_escape(algorithm)}</text>')

    for algorithm in algorithms:
        color = COLORS.get(algorithm, "#111827")
        points = []
        for i, budget in enumerate(budgets):
            row = lookup.get((budget, algorithm))
            if not row:
                continue
            x = x_for(i)
            y = y_for(float(row[metric]))
            points.append((x, y, row))
        if len(points) >= 2:
            point_text = " ".join(f"{x:.2f},{y:.2f}" for x, y, _row in points)
            lines.append(f'<polyline fill="none" stroke="{color}" stroke-width="2.4" points="{point_text}"/>')
        std_key = metric.replace("mean_", "std_")
        for x, y, row in points:
            if std_key in row:
                err = float(row[std_key])
                y_hi = y_for(float(row[metric]) + err)
                y_lo = y_for(max(y_min, float(row[metric]) - err))
                lines.append(f'<line x1="{x:.2f}" y1="{y_hi:.2f}" x2="{x:.2f}" y2="{y_lo:.2f}" stroke="{color}" stroke-width="1.2" opacity="0.7"/>')
                lines.append(f'<line x1="{x - 5:.2f}" y1="{y_hi:.2f}" x2="{x + 5:.2f}" y2="{y_hi:.2f}" stroke="{color}" stroke-width="1.2" opacity="0.7"/>')
                lines.append(f'<line x1="{x - 5:.2f}" y1="{y_lo:.2f}" x2="{x + 5:.2f}" y2="{y_lo:.2f}" stroke="{color}" stroke-width="1.2" opacity="0.7"/>')
            lines.append(f'<circle cx="{x:.2f}" cy="{y:.2f}" r="4.5" fill="{color}" stroke="#ffffff" stroke-width="1.4"/>')

    lines.append("</svg>")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")


def best_by_budget(summary_rows):
    result = []
    grouped = defaultdict(list)
    for row in summary_rows:
        grouped[int(row["budget_miles"])].append(row)
    for budget in sorted(grouped):
        best = max(grouped[budget], key=lambda row: float(row["mean_prize"]))
        result.append((budget, best["algorithm"], float(best["mean_prize"])))
    return result


def write_markdown(path, summary_rows, manifest):
    lines = []
    lines.append("# BC-TSP Comparison Deliverable Summary")
    lines.append("")
    lines.append(f"Generated: {manifest['generated_at']}")
    lines.append("")
    lines.append("## Scope")
    lines.append("")
    lines.append("This deliverable compares Stanley's Attention-based routing workflow against Chris Gonzalez's implemented Greedy 1, Greedy 2, and P-MARL baselines on the same 48-city BC-TSP capital-city instances.")
    lines.append("")
    lines.append("Ant-Q and ILP are intentionally excluded because Dr. Tang's direct request named Greedy 1, Greedy 2, and P-MARL only.")
    lines.append("")
    lines.append("## Experiment Setup")
    lines.append("")
    lines.append("- Dataset: `Capital_Cities.txt` from `source/java-bctsp/src`.")
    lines.append("- Instances: first 20 depot cities used by Chris's `TableData.generateRandomCities` convention.")
    lines.append("- Budgets: 4000, 6000, 8000, and 10000 miles.")
    lines.append("- Metric used for comparison: collected prize excluding depot prize.")
    lines.append("- Distance convention: Java `CityNode.java` Haversine miles, using `6371 km * 0.62 miles/km`.")
    lines.append("- Attention model: pretrained `op_dist_50`, decoded greedily unless the run script is changed.")
    lines.append(f"- Attention feasibility note: {manifest['attention_route_repairs']} of {manifest['attention_rows']} raw Attention routes were trimmed after decoding so the final reported route satisfies the same Java-mile budget; final infeasible rows: {manifest['attention_final_infeasible_rows']}.")
    lines.append("- Reproducibility note: Chris's P-MARL implementation uses unseeded Java randomness during learning, so rerunning the Java wrapper can produce slightly different P-MARL rows.")
    lines.append("")
    lines.append("## Summary Table")
    lines.append("")
    lines.append("| Budget | Algorithm | n | Mean prize | Std prize | Mean distance | Mean runtime ms |")
    lines.append("|---:|---|---:|---:|---:|---:|---:|")
    for row in summary_rows:
        lines.append(
            f"| {row['budget_miles']} | {row['algorithm']} | {row['n']} | "
            f"{float(row['mean_prize']):.1f} | {float(row['std_prize']):.1f} | "
            f"{float(row['mean_distance_miles']):.1f} | {float(row['mean_runtime_ms']):.1f} |"
        )
    lines.append("")
    lines.append("## Best Mean Prize By Budget")
    lines.append("")
    for budget, algorithm, prize in best_by_budget(summary_rows):
        lines.append(f"- {budget} miles: {algorithm} with mean prize {prize:.1f}.")
    lines.append("")
    lines.append("## Reviewer-Scale Note")
    lines.append("")
    lines.append("The reviewer mentioned large-scale evidence around 10,000 nodes. The attached code/data supports the 48-city and 10-city capital datasets, but no 10,000-node BC-TSP instance generator or benchmark dataset was attached. This should be raised as a next-step experiment rather than claimed as completed.")
    lines.append("")
    lines.append("## Files")
    lines.append("")
    lines.append("- `summary/comparison_raw_combined.csv`: all per-instance raw results.")
    lines.append("- `summary/comparison_summary_by_budget.csv`: mean and standard deviation table.")
    lines.append("- `summary/comparison_results.xlsx`: Excel workbook with raw results, summary, and manifest.")
    lines.append("- `figures/prize_vs_budget.svg`, `figures/distance_vs_budget.svg`, `figures/runtime_vs_budget.svg`: plots.")
    path.write_text("\n".join(lines), encoding="utf-8")


def write_manifest(path, manifest):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")


def main():
    args = build_parser().parse_args()
    root = Path(args.root)
    deliverables = root
    raw_rows = []
    raw_rows.extend(read_rows(Path(args.java_csv)))
    raw_rows.extend(read_rows(Path(args.attention_csv)))
    raw_rows.sort(key=lambda row: (row["budget_miles"], row["instance_index"], algorithm_rank(row["algorithm"])))

    summary_rows = summarize(raw_rows)
    manifest = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "dataset": "Capital_Cities.txt",
        "instance_count": len({(row["budget_miles"], row["instance_index"]) for row in raw_rows}),
        "depot_count": len({row["instance_index"] for row in raw_rows}),
        "budgets_miles": sorted({row["budget_miles"] for row in raw_rows}),
        "algorithms": [alg for alg in ALGORITHM_ORDER if any(row["algorithm"] == alg for row in raw_rows)],
        "excluded_algorithms": ["Ant-Q", "ILP"],
        "reason_excluded": "Dr. Tang's direct request named Greedy 1, Greedy 2, and P-MARL; Ant-Q is not implemented in Chris's handoff code and ILP requires external CPLEX.",
        "prize_metric": "collected prize excluding depot prize",
        "distance_metric": "route distance in miles using the CityNode.java convention",
        "attention_rows": sum(1 for row in raw_rows if row["algorithm"] == "Attention"),
        "attention_route_repairs": sum(1 for row in raw_rows if row["algorithm"] == "Attention" and row.get("repair_applied") == "true"),
        "attention_final_infeasible_rows": sum(1 for row in raw_rows if row["algorithm"] == "Attention" and row.get("feasible_under_java_miles") not in ("", "true")),
        "p_marl_reproducibility_note": "Chris's P-MARL implementation uses unseeded Java randomness during learning, so rerunning the Java wrapper can produce slightly different P-MARL rows.",
    }

    write_csv(deliverables / "summary" / "comparison_raw_combined.csv", raw_rows, FIELDNAMES)
    write_csv(deliverables / "summary" / "comparison_summary_by_budget.csv", summary_rows, SUMMARY_FIELDS)
    write_manifest(deliverables / "summary" / "experiment_manifest.json", manifest)
    write_workbook(deliverables / "summary" / "comparison_results.xlsx", raw_rows, summary_rows, manifest)
    write_markdown(deliverables / "summary" / "professor_summary.md", summary_rows, manifest)

    plot_svg(deliverables / "figures" / "prize_vs_budget.svg", summary_rows, "mean_prize", "Collected Prize vs Budget", "Mean collected prize")
    plot_svg(deliverables / "figures" / "distance_vs_budget.svg", summary_rows, "mean_distance_miles", "Route Distance vs Budget", "Mean distance (miles)")
    plot_svg(deliverables / "figures" / "runtime_vs_budget.svg", summary_rows, "mean_runtime_ms", "Runtime vs Budget", "Mean runtime (ms)")
    print(f"Wrote deliverables under {deliverables}")


if __name__ == "__main__":
    main()
